"""
tools/excel_builder.py — Project Antigravity
Canonical Excel builder: 5-tab openpyxl workbook matching Section 8 visual spec.

Tabs:
  1. 📋 Ad Intelligence Records (13 cols, row height 320, freeze D4)
  2. 🎬 Production Formulas (7 cols, row height 280, freeze D3)
  3. ⚡ Key Takeaways (6 cols, row height 220, freeze D3)
  4. 📖 Legend & Instructions (2 cols)
  5. 🏆 Strategic Summary (conditional on summary data)

Input:  records: list[dict] — list of AdRecord dicts
Output: .xlsx file at output_path
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_BG      = "1A1A2E"
PANEL_BG     = "0F3460"
ACCENT_GREEN = "00C896"
ROW_ALT      = "F8F9FA"
ROW_WHITE    = "FFFFFF"
BORDER_COL   = "DEE2E6"

H_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=9, color="212529")
LINK_FONT = Font(name="Arial", size=9, color="0563C1", underline="single")
TITLE_FONT_LG = Font(name="Arial", bold=True, size=13, color="FFFFFF")
TITLE_FONT_MD = Font(name="Arial", bold=True, size=12, color="FFFFFF")
TITLE_FONT_XL = Font(name="Arial", bold=True, size=14, color="FFFFFF")
SUB_FONT   = Font(name="Arial", size=9, color="AAAAAA", italic=True)


def _border():
    s = Side(border_style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)


def _hfill(color=DARK_BG):
    return PatternFill("solid", fgColor=color)


def _rfill(alt=False):
    return PatternFill("solid", fgColor=ROW_ALT if alt else ROW_WHITE)


def _wrap():
    return Alignment(wrap_text=True, vertical="top")


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Field extractors ──────────────────────────────────────────────────────────

def extract_hook_type(hook: str) -> str:
    """Extract hook type label from hook field."""
    m = re.search(r"Type:\s*([^\n]+)", hook)
    return m.group(1).strip() if m else "—"


def extract_big_idea(concept: str) -> str:
    """Extract Big Idea from concept field."""
    m = re.search(r'Big Idea:\s*["\']?([^"\'\n.]+)', concept)
    return m.group(1).strip() if m else "—"


def extract_framework(script: str) -> str:
    """Extract framework name from scriptBreakdown field."""
    m = re.search(r"Framework:\s*([^\n]+)", script)
    return m.group(1).strip() if m else "—"


def parse_takeaways(kt: str) -> tuple:
    """
    Parse keyTakeaways into (steal, kaizen, upgrade) strings.
    Uses ✅/🔨/🚀 emoji markers as delimiters.
    """
    steals  = re.findall(r"✅ STEAL[^:]*:\s*(.*?)(?=\n\n[✅🔨🚀]|\Z)", kt, re.DOTALL)
    kaizens = re.findall(r"🔨 KAIZEN[^:]*:\s*(.*?)(?=\n\n[✅🔨🚀]|\Z)", kt, re.DOTALL)
    upgrade = re.findall(r"🚀 UPGRADE[^:]*:\s*(.*?)(?=\n\n[✅🔨🚀]|\Z)", kt, re.DOTALL)
    steal_txt   = "\n\n".join(f"• {s.strip()}" for s in steals) if steals else "—"
    kaizen_txt  = "\n\n".join(f"• {k.strip()}" for k in kaizens) if kaizens else "—"
    upgrade_txt = upgrade[0].strip() if upgrade else "—"
    return steal_txt, kaizen_txt, upgrade_txt


# ── Helpers ───────────────────────────────────────────────────────────────────

def _col_headers(ws, row, headers_widths, fill_color=PANEL_BG):
    """Write column headers with styling."""
    for col_idx, (label, width) in enumerate(headers_widths, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = H_FONT
        cell.fill = _hfill(fill_color)
        cell.alignment = _center()
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 36


def _title_row(ws, row, text, ncols, font=None, height=32):
    """Write merged title row."""
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = font or TITLE_FONT_LG
    c.fill = _hfill(DARK_BG)
    c.alignment = _center()
    ws.row_dimensions[row].height = height


# ── MAIN BUILDER ──────────────────────────────────────────────────────────────

def build_excel(records: list, output_path: str, your_brand: str = "FusiForce", your_parent: str = "Wellness Nest", summary: dict = None) -> None:
    """
    Build the 5-tab Excel intelligence file.

    Args:
        records: List of AdRecord dicts (must have adScore computed)
        output_path: Path to save .xlsx file
        your_brand: Brand name for title bars
        your_parent: Parent company name for title bars
        summary: Strategic summary dict from Sonnet (optional)
    """
    wb = Workbook()

    # Sort records by adScore DESC for Tab 1
    records_sorted = sorted(records, key=lambda r: r.get("adScore", 0), reverse=True)

    # ════════════════════════════════════════════════════════════
    # TAB 1 — 📋 Ad Intelligence Records
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📋 Ad Intelligence Records"
    ws1.sheet_view.showGridLines = False

    COLS1 = [
        ("#", 5), ("BRAND", 18), ("MARKET", 9),
        ("AD LINK (Foreplay)", 22), ("LANDING PAGE", 22), ("VIDEO URL", 22),
        ("AD START DATE", 14), ("LONGEVITY (DAYS)", 14), ("AD ITERATIONS", 12),
        ("STATUS", 9), ("DURATION (SEC)", 12), ("VIDEO FORMAT", 12),
        ("IMPRESSIONS (LOW)", 15), ("IMPRESSIONS (HIGH)", 15),
        ("SPEND (LOW)", 12), ("SPEND (HIGH)", 12), ("CURRENCY", 9),
        ("AD SCORE", 10),
        ("HOOK", 52), ("CONCEPT / BIG IDEA", 52),
        ("SCRIPT BREAKDOWN", 52), ("VISUAL — A/B/C ROLL", 52),
        ("CONSUMER PSYCHOLOGY", 52), ("CTA", 38),
        ("KEY TAKEAWAYS", 60), ("🎬 PRODUCTION FORMULA", 70),
        ("HOOK TYPE", 28), ("PRIMARY ANGLE", 28), ("FRAMEWORK", 28), ("CREATIVE PATTERN", 24),
        ("PAGE NAME", 18), ("AD LIBRARY ID", 18), ("CRAWLED AT", 16),
    ]

    # Title row 1
    _title_row(
        ws1, 1,
        f"🔬 PROJECT ANTIGRAVITY — COMPETITOR AD INTELLIGENCE DATABASE  |  {your_brand} / {your_parent}",
        len(COLS1), font=TITLE_FONT_LG, height=32
    )

    # Sub-header row 2
    ws1.merge_cells(f"A2:{get_column_letter(len(COLS1))}2")
    sub = ws1["A2"]
    sub.value = (
        "Keyword: Creatine Gummies  |  Markets: US · UK · AU  |  "
        "Framework: Hook → Concept → Script → Visuals → Psychology → CTA → Takeaways → Production Formula"
    )
    sub.font = SUB_FONT
    sub.fill = _hfill("0D0D1F")
    sub.alignment = _center()
    ws1.row_dimensions[2].height = 18

    # Column headers row 3
    _col_headers(ws1, 3, COLS1)

    # Data rows starting at row 4
    for row_i, rec in enumerate(records_sorted):
        r = row_i + 4
        alt = row_i % 2 == 1
        vals = [
            row_i + 1,
            rec.get("brand", ""),
            rec.get("market", rec.get("region", "")),
            rec.get("link_ads", rec.get("foreplayUrl", "")),
            rec.get("link_landing_page", rec.get("landingPageUrl", "")),
            rec.get("videoUrl", ""),
            rec.get("adStartDate", ""),
            rec.get("longevityDays", 0),
            rec.get("adIterationCount", 1),
            "Active" if rec.get("isActive", True) else "Inactive",
            rec.get("durationSeconds", 0),
            rec.get("videoFormat", ""),
            rec.get("impressionsLower", ""),
            rec.get("impressionsUpper", ""),
            rec.get("spendLower", ""),
            rec.get("spendUpper", ""),
            rec.get("spendCurrency", ""),
            rec.get("adScore", 0),
            rec.get("hook", ""),
            rec.get("concept", ""),
            rec.get("script_breakdown", rec.get("scriptBreakdown", "")),
            rec.get("visual_rolls", rec.get("visual", "")),
            rec.get("psychology", ""),
            rec.get("cta", ""),
            rec.get("key_takeaways", rec.get("keyTakeaways", "")),
            rec.get("production_formula", rec.get("productionFormula", "")),
            rec.get("hookType", ""),
            rec.get("primaryAngle", ""),
            rec.get("frameworkName", ""),
            rec.get("creativePattern", ""),
            rec.get("pageName", ""),
            rec.get("adLibraryId", ""),
            rec.get("crawledAt", ""),
        ]
        # Link columns: AD LINK (4), LANDING PAGE (5), VIDEO URL (6)
        link_cols = {4, 5, 6}
        # Centered short columns: #(1), BRAND(2), MARKET(3), 7-18, 27-30, 31-33
        center_cols = {1, 2, 3, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 27, 28, 29, 30, 31, 32, 33}
        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=col_idx, value=val)
            cell.fill = _rfill(alt)
            cell.border = _border()
            if col_idx in link_cols and val:
                cell.hyperlink = str(val)
                cell.font = LINK_FONT
                cell.alignment = _wrap()
            elif col_idx in center_cols:
                cell.font = Font(name="Arial", bold=(col_idx in {2}), size=9)
                cell.alignment = _center()
            else:
                cell.font = BODY_FONT
                cell.alignment = _wrap()
        ws1.row_dimensions[r].height = 320

    ws1.freeze_panes = "D4"
    last_col1 = get_column_letter(len(COLS1))
    ws1.auto_filter.ref = f"A3:{last_col1}3"

    # ════════════════════════════════════════════════════════════
    # TAB 2 — 🎬 Production Formulas
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("🎬 Production Formulas")
    ws2.sheet_view.showGridLines = False

    COLS2 = [
        ("#", 5), ("BRAND", 18), ("MARKET", 9),
        ("HOOK TYPE", 30), ("PRIMARY ANGLE / BIG IDEA", 40),
        ("FRAMEWORK", 35),
        ("⭐ FULL PRODUCTION FORMULA (Phase-by-Phase Shoot Brief)", 90),
    ]

    _title_row(
        ws2, 1,
        f"🎬 PRODUCTION FORMULAS — Ready-to-Brief Scripts for {your_brand} Creative Team",
        len(COLS2), font=TITLE_FONT_MD, height=28
    )
    _col_headers(ws2, 2, COLS2)

    for row_i, rec in enumerate(records_sorted):
        r = row_i + 3
        alt = row_i % 2 == 1
        hook_text = rec.get("hook", "")
        concept_text = rec.get("concept", "")
        script_text = rec.get("script_breakdown", rec.get("scriptBreakdown", ""))
        formula_text = rec.get("production_formula", rec.get("productionFormula", ""))

        vals = [
            row_i + 1,
            rec.get("brand", ""),
            rec.get("market", rec.get("region", "")),
            extract_hook_type(hook_text),
            extract_big_idea(concept_text),
            extract_framework(script_text),
            formula_text,
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=col_idx, value=val)
            cell.fill = _rfill(alt)
            cell.border = _border()
            cell.font = BODY_FONT
            cell.alignment = _wrap() if col_idx > 3 else _center()
        ws2.row_dimensions[r].height = 280

    ws2.freeze_panes = "D3"
    last_col2 = get_column_letter(len(COLS2))
    ws2.auto_filter.ref = f"A2:{last_col2}2"

    # ════════════════════════════════════════════════════════════
    # TAB 3 — ⚡ Key Takeaways
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("⚡ Key Takeaways")
    ws3.sheet_view.showGridLines = False

    COLS3 = [
        ("#", 5), ("BRAND", 18), ("MARKET", 9),
        ("✅ STEAL (What to replicate)", 65),
        ("🔨 KAIZEN (Gap to exploit)", 65),
        ("🚀 UPGRADE (FusiForce advantage)", 65),
    ]

    _title_row(
        ws3, 1,
        f"⚡ KEY TAKEAWAYS — STEAL · KAIZEN · UPGRADE  |  Filtered for {your_brand} Implementation",
        len(COLS3), font=TITLE_FONT_MD, height=28
    )
    _col_headers(ws3, 2, COLS3)

    for row_i, rec in enumerate(records_sorted):
        r = row_i + 3
        alt = row_i % 2 == 1
        kt = rec.get("key_takeaways", rec.get("keyTakeaways", ""))
        steal, kaizen, upgrade = parse_takeaways(kt)
        vals = [
            row_i + 1,
            rec.get("brand", ""),
            rec.get("market", rec.get("region", "")),
            steal, kaizen, upgrade,
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=col_idx, value=val)
            cell.fill = _rfill(alt)
            cell.border = _border()
            cell.font = BODY_FONT
            cell.alignment = _wrap() if col_idx > 3 else _center()
        ws3.row_dimensions[r].height = 220

    ws3.freeze_panes = "D3"
    last_col3 = get_column_letter(len(COLS3))
    ws3.auto_filter.ref = f"A2:{last_col3}2"

    # ════════════════════════════════════════════════════════════
    # TAB 4 — 📖 Legend & Instructions
    # ════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📖 Legend & Instructions")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 80

    _title_row(ws4, 1, "📖 PROJECT ANTIGRAVITY — LEGEND & INSTRUCTIONS", 2,
               font=TITLE_FONT_XL, height=40)

    legend = [
        ("HOW TO USE THIS FILE", ""),
        ("Tab 1 → 📋 Ad Intelligence Records",
         "Complete forensic analysis. All 9 fields per ad. Use for deep-dive research on any specific ad."),
        ("Tab 2 → 🎬 Production Formulas",
         "Phase-by-phase shoot briefs. Filter by Hook Type or Angle. Paste directly into creative brief template."),
        ("Tab 3 → ⚡ Key Takeaways",
         "STEAL · KAIZEN · UPGRADE pre-parsed. Filter by Brand or Market to cluster competitor gaps."),
        ("Tab 4 → This Page",
         "Field definitions, usage guide, scoring explanation."),
        ("", ""),
        ("FIELD DEFINITIONS", ""),
        ("Hook",
         "Hook TYPE label + exact execution (what happens 0–5s on screen) + WHY it stops the scroll (psychological mechanism named)."),
        ("Concept / Big Idea",
         "Central creative idea in one sentence + full strategic architecture + secondary angles (bulleted)."),
        ("Script Breakdown",
         "Named copywriting framework + numbered beats with exact timecodes. Shows the full narrative arc."),
        ("Visual — A/B/C Roll",
         "A-Roll = presenter/main footage. B-Roll = product/prop shots with timecodes. C-Roll = text overlays, certs, reviews on screen (or strategic note if absent)."),
        ("Consumer Psychology",
         "Named cognitive biases with exact execution detail per bias. Includes regional market resonance note (US/UK/AU)."),
        ("CTA",
         "Mechanism (verbal/gesture/text/button) + offer shown or why no offer + what the landing page must handle."),
        ("Key Takeaways",
         "✅ STEAL = replicate with FusiForce implementation instructions.\n"
         "🔨 KAIZEN = structural weakness + how FusiForce exploits it.\n"
         "🚀 UPGRADE = where FusiForce structurally wins vs. this brand."),
        ("Production Formula",
         "Ready-to-shoot brief: FORMAT line + 5 phases (HOOK/AGITATE/REVEAL/TRUST/CTA).\n"
         "Each phase: screen direction + 📝 voiceover line + 🖥 TEXT SUPER."),
        ("", ""),
        ("AD SCORING", ""),
        ("Ad Score (0–10)",
         "Composite: Longevity 40% + Impressions 25% + Iterations 25% + Duration 10%. Data-driven — AI never scores quality."),
        ("Longevity",
         "Days since ad_delivery_start_time. 90+ days = ROI-positive signal. Highest weight metric. "
         "Caveat: VC-backed brands may run unprofitable ads for awareness — treat high longevity from funded brands with additional scrutiny."),
        ("Impressions note",
         "AdScore uses upper bound of Meta's impression range (Meta only provides ranges, not exact numbers). "
         "Ads with wider ranges (e.g., 100K-2M) carry more uncertainty than narrow ranges (e.g., 900K-1.1M). "
         "Use longevity and iteration count as tiebreakers when impression ranges are wide."),
        ("Creative Pattern",
         "One of: Problem-First UGC | Result-First Scroll Stop | Curiosity Gap | Social Proof Cascade | Comparison/Versus | Authority Demo. Classified by Sonnet (descriptive only)."),
        ("", ""),
        ("YOUR BRAND", ""),
        (f"{your_brand} / {your_parent}",
         f"All STEAL/KAIZEN/UPGRADE items and Production Formulas are written specifically for {your_brand} implementation."),
    ]

    SECTION_KEYS = {"HOW TO USE THIS FILE", "FIELD DEFINITIONS", "AD SCORING", "YOUR BRAND"}
    for row_i, (key, val) in enumerate(legend, 2):
        ka = ws4.cell(row=row_i, column=1, value=key)
        va = ws4.cell(row=row_i, column=2, value=val)
        if key in SECTION_KEYS:
            ka.font = Font(name="Arial", bold=True, size=10, color=DARK_BG)
            ka.fill = PatternFill("solid", fgColor="E8F4F8")
            va.fill = PatternFill("solid", fgColor="E8F4F8")
        else:
            f = _rfill(row_i % 2 == 0)
            ka.font = Font(name="Arial", bold=True, size=9, color="333333")
            ka.fill = f
            va.fill = f
        va.font = Font(name="Arial", size=9, color="444444")
        va.alignment = Alignment(wrap_text=True, vertical="top")
        ka.alignment = Alignment(wrap_text=True, vertical="top")
        ka.border = _border()
        va.border = _border()
        ws4.row_dimensions[row_i].height = 40

    # ════════════════════════════════════════════════════════════
    # CONDITIONAL FORMATTING — Green highlight for Top 5 in Tab 1
    # ════════════════════════════════════════════════════════════
    TOP5_FILL = PatternFill("solid", fgColor="E6F9F0")
    top5_count = min(5, len(records_sorted))
    for row_i in range(top5_count):
        r = row_i + 4  # Data starts at row 4
        for col_idx in range(1, len(COLS1) + 1):
            cell = ws1.cell(row=r, column=col_idx)
            cell.fill = TOP5_FILL

    # ════════════════════════════════════════════════════════════
    # TAB 5 — 📊 Strategic Summary
    # ════════════════════════════════════════════════════════════
    if summary and any(v != "—" for v in summary.values()):
        ws5 = wb.create_sheet("📊 Strategic Summary")
        ws5.sheet_view.showGridLines = False
        ws5.column_dimensions["A"].width = 30
        ws5.column_dimensions["B"].width = 100

        _title_row(ws5, 1, f"📊 STRATEGIC SUMMARY — Data-Driven Intelligence for {your_brand}", 2,
                   font=TITLE_FONT_XL, height=40)

        summary_sections = [
            ("DOMINANT PATTERNS", summary.get("dominantPatterns", "—")),
            ("", ""),
            ("TOP 5 WINNERS (by AdScore)", summary.get("top5Analysis", "—")),
            ("", ""),
            ("MARKET INSIGHTS", summary.get("marketInsights", "—")),
            ("", ""),
            ("STRATEGIC RECOMMENDATION", summary.get("strategicRecommendation", "—")),
            ("", ""),
            ("COMPETITOR RANKING", summary.get("competitorRanking", "—")),
        ]

        SECTION_KEYS_5 = {"DOMINANT PATTERNS", "TOP 5 WINNERS (by AdScore)", "MARKET INSIGHTS",
                          "STRATEGIC RECOMMENDATION", "COMPETITOR RANKING"}
        for row_i, (key, val) in enumerate(summary_sections, 2):
            ka = ws5.cell(row=row_i, column=1, value=key)
            va = ws5.cell(row=row_i, column=2, value=val)
            if key in SECTION_KEYS_5:
                ka.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
                ka.fill = _hfill(PANEL_BG)
                va.fill = PatternFill("solid", fgColor="F0F7FF")
            else:
                f = _rfill(row_i % 2 == 0)
                ka.font = Font(name="Arial", bold=True, size=9, color="333333")
                ka.fill = f
                va.fill = f
            va.font = Font(name="Arial", size=9, color="333333")
            va.alignment = Alignment(wrap_text=True, vertical="top")
            ka.alignment = Alignment(wrap_text=True, vertical="top")
            ka.border = _border()
            va.border = _border()
            ws5.row_dimensions[row_i].height = 80 if key in SECTION_KEYS_5 else 20

    wb.save(output_path)
    print(f"✅ Excel saved: {output_path}")


# ── CLI entry point ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 3:
        print("Usage: python excel_builder.py <records.json> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)

    build_excel(data, sys.argv[2])
